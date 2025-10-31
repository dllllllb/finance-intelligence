"""
BNK 전체 잔가율 테이블 추출 - 8개 잔가사
"""
import pandas as pd
import openpyxl
import json

def extract_all_rv_tables():
    """BNK RVs 시트에서 모든 잔가사 잔가율 테이블 추출"""
    filepath = "BNK-25-10-V4.xlsm"

    print("="*80)
    print("📊 BNK 전체 잔가율 테이블 추출 (8개 잔가사)")
    print("="*80)

    # openpyxl로 값만 읽기
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb['RVs']

    rv_tables = {}

    # 잔가사별 테이블 위치 정의
    rv_configs = [
        {'name': '웨스트_통합', 'start_row': 4, 'grades': ['S', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']},
        {'name': '웨스트_수입', 'start_row': 17, 'grades': ['S', 'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']},
        {'name': '큐브_수입', 'start_row': 30, 'grades': ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']},
        {'name': '무카_국산', 'start_row': 41, 'grades': ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P']},
        {'name': '태양_수입', 'start_row': 53, 'grades': ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']},
        {'name': '조이_수입', 'start_row': 66, 'grades': ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']},
        {'name': '코렉트', 'start_row': 78, 'grades': ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']},
        {'name': 'ADB', 'start_row': 88, 'grades': ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']},
    ]

    periods = [12, 24, 36, 42, 44, 48, 60]

    for config in rv_configs:
        print(f"\n{config['name']} 추출 중...")
        table_name = f"{config['name']}_2만"
        rv_table = {}

        for i, period in enumerate(periods):
            row_idx = config['start_row'] + i + 1  # openpyxl은 1-based
            rv_table[period] = {}

            for j, grade in enumerate(config['grades']):
                col_idx = 1 + j + 1  # openpyxl은 1-based
                value = ws.cell(row_idx, col_idx).value
                if value is not None and isinstance(value, (int, float)) and value != 0:
                    rv_table[period][grade] = float(value)

        rv_tables[table_name] = rv_table
        print(f"  ✓ {len(rv_table)} 기간, {len(config['grades'])} 등급")

    # 감가율 테이블 (개월수별)
    print("\n개월수별 감가율 테이블 추출 중...")
    depreciation_table = {}
    for month in range(1, 61):
        row_idx = 6 + month + 1  # openpyxl은 1-based
        col_idx = 32 + 1  # openpyxl은 1-based
        value = ws.cell(row_idx, col_idx).value
        if value is not None and isinstance(value, (int, float)):
            depreciation_table[month] = float(value)

    rv_tables['감가율_테이블'] = depreciation_table
    print(f"  ✓ {len(depreciation_table)} 개월")

    # 주행거리 조정
    rv_tables['주행거리_조정'] = {
        '1만': +0.02,
        '1.5만': +0.01,
        '2만': 0,
        '3만': -0.03
    }

    # JSON 저장
    output_path = "src/bnk_rv_tables.json"
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(rv_tables, f, ensure_ascii=False, indent=2)

    print(f"\n✅ 저장 완료: {output_path}")
    print(f"총 {len(rv_tables)} 테이블 (8개 잔가사 + 감가율 + 주행거리 조정)")

    # 샘플 출력
    print("\n📋 샘플 데이터:")
    for table_name in ['웨스트_통합_2만', '태양_수입_2만', '조이_수입_2만']:
        if table_name in rv_tables and 36 in rv_tables[table_name]:
            grades = list(rv_tables[table_name][36].keys())
            if grades:
                first_grade = grades[0]
                print(f"  {table_name} 36개월 {first_grade}등급: {rv_tables[table_name][36][first_grade]}")

    print("\n" + "="*80)

if __name__ == "__main__":
    extract_all_rv_tables()
