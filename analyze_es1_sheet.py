"""
Es1 시트 분석 - 실제 계산 로직 확인
"""
import pandas as pd
import openpyxl

def analyze_es1_sheet():
    """Es1 시트 상세 분석"""
    filepath = "BNK-25-10-V4.xlsm"

    print("="*80)
    print("🔍 Es1 시트 분석 - 실제 계산 로직")
    print("="*80)

    try:
        wb = openpyxl.load_workbook(filepath, data_only=False)
        ws = wb['Es1']

        # 주요 셀 분석
        cells_to_check = [
            'B26', 'B30', 'B42', 'B46', 'B48', 'B54', 'B56', 'B62', 'B70',
            'B83', 'B133', 'B134', 'B136', 'B139', 'B382'
        ]

        print("\n주요 셀 값과 수식:")
        print("-"*80)
        for cell_addr in cells_to_check:
            cell = ws[cell_addr]
            # 주변 레이블 찾기 (왼쪽 셀)
            col_letter = cell_addr[0]
            row_num = int(cell_addr[1:])
            label_cell = ws[f'A{row_num}']
            label = label_cell.value if label_cell.value else ""

            print(f"\n[{cell_addr}] {label}")
            if cell.value:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    print(f"  수식: {cell.value}")
                else:
                    print(f"  값: {cell.value}")
            else:
                print(f"  (비어있음)")

        # 취득세 관련 (B26, B133, B134)
        print("\n\n=== 취득세 / 취득원가 관련 셀 ===")
        print("-"*80)
        for row in range(20, 150):
            label_cell = ws[f'A{row}']
            if label_cell.value and isinstance(label_cell.value, str):
                if any(keyword in label_cell.value for keyword in ['취득세', '취득원가', '공채', '등록세']):
                    value_cell = ws[f'B{row}']
                    print(f"\n[A{row}] {label_cell.value}")
                    print(f"[B{row}] ", end="")
                    if value_cell.value:
                        if isinstance(value_cell.value, str) and value_cell.value.startswith('='):
                            print(f"수식: {value_cell.value}")
                        else:
                            print(f"값: {value_cell.value}")
                    else:
                        print("(비어있음)")

        # 잔가 관련 (B56, B139)
        print("\n\n=== 잔가 / 잔존가치 관련 셀 ===")
        print("-"*80)
        for row in range(40, 150):
            label_cell = ws[f'A{row}']
            if label_cell.value and isinstance(label_cell.value, str):
                if any(keyword in label_cell.value for keyword in ['잔가', '잔존', 'RV', '고잔가']):
                    value_cell = ws[f'B{row}']
                    print(f"\n[A{row}] {label_cell.value}")
                    print(f"[B{row}] ", end="")
                    if value_cell.value:
                        if isinstance(value_cell.value, str) and value_cell.value.startswith('='):
                            print(f"수식: {value_cell.value}")
                        else:
                            print(f"값: {value_cell.value}")
                    else:
                        print("(비어있음)")

    except Exception as e:
        print(f"오류: {e}")
        import traceback
        traceback.print_exc()

    # RVs 시트 전체 구조 보기
    print("\n\n=== RVs 시트 전체 구조 ===")
    print("-"*80)
    try:
        df_rv = pd.read_excel(filepath, sheet_name='RVs', engine='openpyxl', header=None)

        print(f"\n총 행 수: {len(df_rv)}, 총 열 수: {len(df_rv.columns)}")

        # 테이블 구조 파악 (헤더 찾기)
        print("\n테이블 헤더들:")
        for i in range(min(100, len(df_rv))):
            first_col = str(df_rv.iloc[i, 0])
            if first_col and not first_col.isdigit() and len(first_col) > 2:
                print(f"  [row {i}] {first_col}", end="")
                if len(df_rv.columns) > 1 and pd.notna(df_rv.iloc[i, 1]):
                    print(f" - {df_rv.iloc[i, 1]}")
                else:
                    print()

    except Exception as e:
        print(f"오류: {e}")

    print("\n" + "="*80)

if __name__ == "__main__":
    analyze_es1_sheet()
