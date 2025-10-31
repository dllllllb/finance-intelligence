"""
BNK 엑셀 계산 로직 완전 분석 - 상세버전
"""
import pandas as pd
import openpyxl

def analyze_detailed_logic():
    """BNK 견적 계산 로직 상세 분석"""
    filepath = "BNK-25-10-V4.xlsm"

    print("="*80)
    print("🔍 BNK 견적서 계산 로직 상세 분석")
    print("="*80)

    # openpyxl로 수식 분석
    print("\n📋 운용리스견적 시트 - 수식 분석")
    print("-"*80)
    try:
        wb = openpyxl.load_workbook(filepath, data_only=False)
        ws = wb['운용리스견적']

        # 주요 셀 찾기
        print("\n입력값 찾기:")
        for row in range(1, 100):
            for col in range(1, 30):
                cell = ws.cell(row, col)
                if cell.value and isinstance(cell.value, str):
                    if any(keyword in cell.value for keyword in ['차량가', '취득세', '취득원가', '잔가', '잔존', '계약기간', '주행']):
                        print(f"  {cell.coordinate}: {cell.value}")
                        # 옆 셀 확인 (값이나 수식)
                        next_cell = ws.cell(row, col+1)
                        if next_cell.value:
                            if isinstance(next_cell.value, str) and next_cell.value.startswith('='):
                                print(f"    → 수식: {next_cell.value}")
                            else:
                                print(f"    → 값: {next_cell.value}")

        print("\n\n계산 관련 주요 수식:")
        keywords = ['취득세', '취득원가', '잔가', '잔존', '월대여료', '월납입']
        for row in range(1, 100):
            for col in range(1, 30):
                cell = ws.cell(row, col)
                if cell.value and isinstance(cell.value, str):
                    # 수식인 경우
                    if cell.value.startswith('='):
                        # 앞 셀 확인 (레이블)
                        label_cell = ws.cell(row, col-1)
                        label = str(label_cell.value) if label_cell.value else ""

                        if any(keyword in label for keyword in keywords):
                            print(f"\n  [{cell.coordinate}] {label}")
                            print(f"    수식: {cell.value}")

    except Exception as e:
        print(f"오류: {e}")

    # RVs 시트에서 BMW X5 찾기
    print("\n\n📋 RVs 시트 - BMW X5 잔가율 찾기")
    print("-"*80)
    try:
        df_rv = pd.read_excel(filepath, sheet_name='RVs', engine='openpyxl', header=None)

        print("\nRVs 시트 구조 (처음 20행, 20열):")
        for i in range(min(20, len(df_rv))):
            row_data = []
            for j in range(min(20, len(df_rv.columns))):
                val = df_rv.iloc[i, j]
                if pd.notna(val):
                    row_data.append(f"[{i},{j}]: {val}")
            if row_data:
                print("  " + " | ".join(row_data))

        # BMW, X5 검색
        print("\n\nBMW 또는 X5 검색:")
        for i in range(len(df_rv)):
            for j in range(len(df_rv.columns)):
                val = str(df_rv.iloc[i, j])
                if 'BMW' in val.upper() or 'X5' in val.upper():
                    print(f"  [{i},{j}]: {val}")

    except Exception as e:
        print(f"오류: {e}")

    # 할부기준 시트 분석
    print("\n\n📋 할부기준 시트 - 취득세율 찾기")
    print("-"*80)
    try:
        df_할부 = pd.read_excel(filepath, sheet_name='할부기준', engine='openpyxl', header=None)

        print("\n취득세 관련 내용:")
        for i in range(min(50, len(df_할부))):
            for j in range(min(20, len(df_할부.columns))):
                val = str(df_할부.iloc[i, j])
                if '취득세' in val or '세율' in val:
                    print(f"  [{i},{j}]: {val}")
                    # 주변 셀도 확인
                    if j+1 < len(df_할부.columns):
                        next_val = df_할부.iloc[i, j+1]
                        if pd.notna(next_val):
                            print(f"    → {next_val}")

    except Exception as e:
        print(f"오류: {e}")

    print("\n" + "="*80)
    print("✅ 분석 완료")
    print("="*80)

if __name__ == "__main__":
    analyze_detailed_logic()
